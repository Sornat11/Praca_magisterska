import os
import re
import logging
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

MODEL_NAME_PATTERN = re.compile(r'Loading model structure and parameters from.*?/(\w+)-')
DATASET_PATTERN = re.compile(r'data_path\s*=\s*.*?/([\w-]+)')
LR_PATTERN = re.compile(r'learning_rate\s*=\s*([\d\.]+)')
EMB_PATTERN = re.compile(r'embedding_size\s*=\s*(\d+)')

TRAIN_TIME_PATTERN = re.compile(r'epoch \d+ training \[time: ([\d\.]+)s')
EVAL_TIME_PATTERN = re.compile(r'epoch \d+ evaluating \[time: ([\d\.]+)s')

SAVE_MODEL_PATTERN = re.compile(r'Saving current: ([\w\/\.-]+\.pth)')
LOAD_MODEL_PATTERN = re.compile(r'Loading model structure and parameters from ([\w\/\.-]+\.pth)')

TEST_RESULT_PATTERN = re.compile(r'test result:\s*OrderedDict\(\{(.*?)\}\)')
BEST_VALID_PATTERN = re.compile(r'best valid\s*:\s*OrderedDict\(\{(.*?)\}\)')
METRIC_KV_PATTERN = re.compile(r"'([\w@]+)':\s*([\d\.]+)")

def parse_log_file(file_path: str) -> Optional[Dict[str, Any]]:
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Podstawowe info
    model_name_match = MODEL_NAME_PATTERN.search(content)
    dataset_match = DATASET_PATTERN.search(content)
    lr_match = LR_PATTERN.search(content)
    emb_match = EMB_PATTERN.search(content)
    
    # Czas
    train_times = TRAIN_TIME_PATTERN.findall(content)
    eval_times = EVAL_TIME_PATTERN.findall(content)
    total_time_min = round((sum(float(t) for t in train_times) + sum(float(t) for t in eval_times)) / 60, 2)

    # Scieżka modelu
    model_path_match = SAVE_MODEL_PATTERN.search(content)
    if not model_path_match:
        model_path_match = LOAD_MODEL_PATTERN.search(content)

    # Wyniki (Testowe lub najlepsze walidacyjne jako fallback)
    metrics = {}
    test_match = TEST_RESULT_PATTERN.search(content)
    
    if test_match:
        raw = test_match.group(1)
        metrics = {f"Test_{p[0].split('@')[0].capitalize()}": float(p[1]) for p in METRIC_KV_PATTERN.findall(raw)}
    else:
        # Jeśli brak testu, bierzemy Best Valid
        best_match = BEST_VALID_PATTERN.search(content)
        if best_match:
            raw = best_match.group(1)
            metrics = {f"Valid_{p[0].split('@')[0].capitalize()}": float(p[1]) for p in METRIC_KV_PATTERN.findall(raw)}

    if not metrics: return None

    file_name = os.path.basename(file_path)
    model_from_file = file_name.split('-')[0]

    data = {
        'Timestamp': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M'),
        'Model': model_name_match.group(1) if model_name_match else model_from_file,
        'Dataset': dataset_match.group(1) if dataset_match else 'Unknown',
        'LR': float(lr_match.group(1)) if lr_match else None,
        'Emb': int(emb_match.group(1)) if emb_match else None,
        'Time (min)': total_time_min,
        'Model Path': model_path_match.group(1) if model_path_match else 'N/A'
    }
    data.update(metrics)
    return data

def aggregate_results(log_dir: str = 'log', output_path: str = '3_Evaluation/experiment_summary') -> None:
    all_results = []
    for root, dirs, files in os.walk(log_dir):
        for file in files:
            if file.endswith('.log'):
                try:
                    res = parse_log_file(os.path.join(root, file))
                    if res: 
                        all_results.append(res)
                except Exception as e:
                    logging.warning(f"Nie udało się sparsować pliku {file}: {e}")

    if not all_results: 
        return
        
    df = pd.DataFrame(all_results).sort_values(by='Timestamp', ascending=False)
    
    # Wybieramy tylko najważniejsze kolumny i zmieniamy ich kolejność
    cols = ['Timestamp', 'Model', 'Dataset', 'LR', 'Emb', 'Time (min)']
    metric_cols = [c for c in df.columns if c.startswith('Test_') or c.startswith('Valid_')]
    cols.extend(sorted(metric_cols))
    cols.append('Model Path')
    
    # Dodanie brakujących kolumn
    missing_cols = [c for c in cols if c not in df.columns]
    for c in missing_cols:
        df[c] = None
        
    df = df[cols]

    if HAS_OPENPYXL:
        excel_file = output_path + '.xlsx'
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Results')
                ws = writer.sheets['Results']
                
                # Prosty styl nagłówka
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    ws.column_dimensions[get_column_letter(col)].width = 15

                # Kolorowanie najlepszego NDCG na zielono
                for col_idx, col_name in enumerate(df.columns, 1):
                    if 'ndcg' in str(col_name).lower():
                        max_val = df[col_name].max()
                        for row_idx, val in enumerate(df[col_name], 2):
                            if val == max_val and val is not None and val > 0:
                                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            print(f"Sukces! Wyniki zapisane w: {excel_file}")
        except PermissionError:
            print("BŁĄD: Zamknij plik Excel!")
    else:
        df.to_csv(output_path + '.csv', index=False)
        print("Brak openpyxl, zapisano CSV.")

if __name__ == '__main__':
    aggregate_results()
